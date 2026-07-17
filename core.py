"""
批量音频响度/统计 - 核心分析逻辑（与 UI 无关）
拖入文件/文件夹 -> 扫描常见音频格式 -> 计算 Peak/RMS/LUFS/LRA/True Peak/DC Offset/削波等 -> 导出 Excel
依赖：系统需安装 ffmpeg / ffprobe 并在 PATH 中可用。
"""

import os
import re
import sys
import json
import shutil
import subprocess
import threading
from pathlib import Path

import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

AUDIO_EXTS = {
    ".wav", ".mp3", ".flac", ".m4a", ".aac", ".ogg", ".oga", ".opus",
    ".wma", ".aiff", ".aif", ".ape", ".dsf", ".wv",
}

# 没有真实位深概念的有损编码（ffprobe 的 sample_fmt 只是解码缓冲区格式，非原始位深）
LOSSY_CODECS = {"mp3", "aac", "vorbis", "wmav1", "wmav2", "wmapro", "opus"}

CLIP_THRESHOLD = 0.999      # 判定接近满幅的采样阈值
CLIP_MIN_CONSECUTIVE = 3    # 连续多少个采样点超阈值才算真正削波（区分单点脉冲噪声）
READ_CHUNK_FRAMES = 1_000_000  # 每次从 ffmpeg 管道读取的帧数（每帧含全部声道）

# Audition "响度(旧版)" 的官方算法未公开，社区反推：50ms 分帧 RMS(dB)，
# 过滤掉 -60dB 以下的静音帧后，对剩余帧的 dB 值取算术平均。这里按同样思路实现，可能与 Audition 数值有 1-2 dB 偏差。
LEGACY_LOUDNESS_WINDOW_MS = 50
LEGACY_LOUDNESS_SILENCE_DB = -60.0

# Windows 下隐藏 ffmpeg/ffprobe 弹出的控制台窗口
_CREATIONFLAGS = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0


class ProcCancelled(Exception):
    """被 CancelToken 中断时抛出，UI 层用来把中断跟正常错误区分开"""


class CancelToken:
    """把当前正在跑的 ffmpeg/ffprobe 进程绑到 token，cancel() 立即 kill，不再等它自然结束。
    workers 每次进入 core 前把 token 传进去，取消时对 token 调 cancel() 就能立刻终止。"""

    def __init__(self):
        self._cancelled = False
        self._proc = None
        self._lock = threading.Lock()

    def cancel(self):
        with self._lock:
            self._cancelled = True
            p = self._proc
        if p is not None:
            try:
                p.kill()
            except Exception:
                pass

    @property
    def cancelled(self):
        return self._cancelled

    def check(self):
        if self._cancelled:
            raise ProcCancelled()

    def _bind(self, proc):
        """把新起的子进程注册到 token；如果已经 cancel 了就立刻 kill 并返回 False"""
        with self._lock:
            if self._cancelled:
                try:
                    proc.kill()
                except Exception:
                    pass
                return False
            self._proc = proc
            return True

    def _unbind(self):
        with self._lock:
            self._proc = None


def _run(cmd, *, cancel_token=None, capture_output=True, stdin=None):
    """subprocess.run 的可取消版本。若 cancel_token 已 cancel，起进程后立刻 kill 并抛 ProcCancelled。"""
    stdout = subprocess.PIPE if capture_output else None
    stderr = subprocess.PIPE if capture_output else None
    proc = subprocess.Popen(
        cmd, stdout=stdout, stderr=stderr, stdin=stdin,
        creationflags=_CREATIONFLAGS,
    )
    if cancel_token is not None:
        if not cancel_token._bind(proc):
            try:
                proc.wait(timeout=2)
            except Exception:
                pass
            raise ProcCancelled()
    try:
        out, err = proc.communicate()
    finally:
        if cancel_token is not None:
            cancel_token._unbind()
    if cancel_token is not None and cancel_token.cancelled:
        raise ProcCancelled()

    class _R:
        pass
    r = _R()
    r.returncode = proc.returncode
    r.stdout = out
    r.stderr = err
    return r


def find_tool(name):
    # PyInstaller 打包后，ffmpeg/ffprobe 会被解压到 sys._MEIPASS 目录，优先用打进 exe 里的那份，
    # 避免依赖用户系统的 PATH 配置
    if getattr(sys, "frozen", False):
        base = Path(getattr(sys, "_MEIPASS", ""))
        candidate = base / f"{name}.exe"
        if candidate.exists():
            return str(candidate)
    return shutil.which(name)


def db(value):
    """线性幅度 -> dBFS，静音返回 -inf"""
    if value is None or value <= 0:
        return float("-inf")
    return 20.0 * np.log10(value)


def fmt_db(value, digits=2, unit=""):
    """默认两位小数并强制补零（-12 -> -12.00），避免同列数值宽度不一致"""
    if value is None:
        return "N/A"
    if value == float("-inf"):
        return f"-inf {unit}" if unit else "-inf"
    text = f"{float(value):.{digits}f}"
    return f"{text} {unit}" if unit else text


def fmt_seconds(seconds):
    if seconds is None:
        return "N/A"
    m, s = divmod(seconds, 60)
    h, m = divmod(int(m), 60)
    if h:
        return f"{h}:{m:02d}:{s:05.2f}"
    return f"{m}:{s:05.2f}"


def probe_file(ffprobe, path, cancel_token=None):
    """用 ffprobe 读取容器/流层面的元数据"""
    cmd = [
        ffprobe, "-v", "quiet", "-print_format", "json",
        "-show_format", "-show_streams", str(path),
    ]
    result = _run(cmd, cancel_token=cancel_token)
    if result.returncode != 0:
        raise RuntimeError(f"ffprobe 失败: {result.stderr.decode(errors='ignore')[:300]}")

    info = json.loads(result.stdout.decode(errors="ignore"))

    audio_stream = None
    for s in info.get("streams", []):
        if s.get("codec_type") == "audio":
            audio_stream = s
            break
    if audio_stream is None:
        raise RuntimeError("未找到音频流")

    sample_rate = int(audio_stream.get("sample_rate", 0) or 0)
    channels = int(audio_stream.get("channels", 0) or 0)
    codec_name = audio_stream.get("codec_name", "")
    sample_fmt = audio_stream.get("sample_fmt", "") or ""

    bit_depth = audio_stream.get("bits_per_raw_sample") or audio_stream.get("bits_per_sample")
    if codec_name in LOSSY_CODECS:
        # 有损编码没有真实位深，ffprobe 报出的 sample_fmt 只是解码缓冲区格式，非原始位深
        bit_depth = None
    elif not bit_depth or int(bit_depth) == 0:
        fallback = {
            "u8": 8, "u8p": 8,
            "s16": 16, "s16p": 16,
            "s32": 32, "s32p": 32,
            "flt": 32, "fltp": 32,
            "dbl": 64, "dblp": 64,
        }
        bit_depth = fallback.get(sample_fmt)

    duration = audio_stream.get("duration") or info.get("format", {}).get("duration")
    duration = float(duration) if duration else None

    # 码率优先取音频流上的 bit_rate；有些容器（如 MP3 CBR）只在 format 里给
    stream_br = audio_stream.get("bit_rate") or info.get("format", {}).get("bit_rate")
    try:
        bit_rate = int(stream_br) if stream_br else None
    except (TypeError, ValueError):
        bit_rate = None

    return {
        "sample_rate": sample_rate,
        "channels": channels,
        "bit_depth": int(bit_depth) if bit_depth else None,
        "duration": duration,
        "codec_name": codec_name,
        "sample_fmt": sample_fmt,
        "bit_rate": bit_rate,
    }


def analyze_basic_stats(ffmpeg, path, channels, cancel_token=None):
    """将音频解码为 float32 PCM，流式计算 Peak / RMS / DC Offset / 削波，避免整段读入内存"""
    if channels <= 0:
        channels = 1
    bytes_per_frame = 4 * channels

    cmd = [
        ffmpeg, "-v", "error", "-i", str(path),
        "-f", "f32le", "-acodec", "pcm_f32le", "pipe:1",
    ]
    proc = subprocess.Popen(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        creationflags=_CREATIONFLAGS,
    )
    # 注册到 cancel_token；已 cancel 就立刻 kill
    if cancel_token is not None and not cancel_token._bind(proc):
        try:
            proc.wait(timeout=2)
        except Exception:
            pass
        raise ProcCancelled()

    count = 0
    sum_ch = np.zeros(channels, dtype=np.float64)
    sumsq_ch = np.zeros(channels, dtype=np.float64)
    peak = 0.0
    clip_count = 0
    tail = np.zeros((0, channels), dtype=np.float32)
    leftover = b""

    read_size = READ_CHUNK_FRAMES * bytes_per_frame
    while True:
        chunk = proc.stdout.read(read_size)
        if not chunk:
            break
        data = leftover + chunk
        usable_len = (len(data) // bytes_per_frame) * bytes_per_frame
        leftover = data[usable_len:]
        if usable_len == 0:
            continue

        arr = np.frombuffer(data[:usable_len], dtype="<f4").reshape(-1, channels)
        count += arr.shape[0]
        sum_ch += arr.sum(axis=0, dtype=np.float64)
        sumsq_ch += (arr.astype(np.float64) ** 2).sum(axis=0)
        chunk_peak = np.max(np.abs(arr)) if arr.size else 0.0
        if chunk_peak > peak:
            peak = float(chunk_peak)

        # 统计削波次数：每一段连续 >= CLIP_MIN_CONSECUTIVE 个满幅采样点算一次削波事件
        check_arr = np.concatenate([tail, arr], axis=0)
        mask = np.abs(check_arr) >= CLIP_THRESHOLD
        if mask.any():
            for c in range(channels):
                col = mask[:, c].view(np.int8)
                if not col.any():
                    continue
                padded = np.concatenate(([0], col, [0]))
                edges = np.flatnonzero(np.diff(padded))
                run_lengths = edges[1::2] - edges[0::2]
                clip_count += int(np.sum(run_lengths >= CLIP_MIN_CONSECUTIVE))
        tail_len = CLIP_MIN_CONSECUTIVE - 1
        tail = arr[-tail_len:] if arr.shape[0] >= tail_len else arr

    proc.stdout.close()
    stderr_data = proc.stderr.read()
    proc.wait()
    if cancel_token is not None:
        cancel_token._unbind()
        if cancel_token.cancelled:
            raise ProcCancelled()
    if proc.returncode != 0 and count == 0:
        raise RuntimeError(f"ffmpeg 解码失败: {stderr_data.decode(errors='ignore')[:300]}")

    if count == 0:
        return {
            "peak_db": float("-inf"),
            "rms_db": [float("-inf")] * channels,
            "dc_offset": 0.0,
            "clip_count": 0,
        }

    mean_ch = sum_ch / count
    rms_ch = np.sqrt(sumsq_ch / count)

    dc_idx = int(np.argmax(np.abs(mean_ch)))
    dc_offset = float(mean_ch[dc_idx])

    return {
        "peak_db": db(peak),
        "rms_db": [db(v) for v in rms_ch],
        "dc_offset": dc_offset,
        "clip_count": clip_count,
    }


_FRAME_RE = re.compile(r"\bM:\s*(-?inf|-?\d+\.?\d*)\s+S:\s*(-?inf|-?\d+\.?\d*)")


def _parse_loudness_value(text):
    return float("-inf") if "inf" in text else float(text)


def analyze_loudness(ffmpeg, path, cancel_token=None):
    """调用 ffmpeg ebur128 滤镜计算 LUFS-I / LUFS-S(max) / LUFS-M(max) / LRA / True Peak"""
    cmd = [
        ffmpeg, "-nostats", "-i", str(path),
        "-filter_complex", "ebur128=peak=true:framelog=info",
        "-f", "null", "-",
    ]
    result = _run(cmd, cancel_token=cancel_token)
    text = result.stderr.decode(errors="ignore")

    m_vals, s_vals = [], []
    for m, s in _FRAME_RE.findall(text):
        m_vals.append(_parse_loudness_value(m))
        s_vals.append(_parse_loudness_value(s))

    def extract(label, pattern):
        # ffmpeg 7.x 的 ebur128 会在开始处理前先打印一次全 0 / -70 / -inf 的初始 Summary，
        # 处理完再打印一次真实值，所以必须取最后一次匹配，否则永远读到假数据
        matches = list(re.finditer(label + r".*?" + pattern, text, re.DOTALL))
        return float(matches[-1].group(1)) if matches else None

    lufs_i = extract(r"Integrated loudness:", r"I:\s*(-?\d+\.?\d*)")
    lra = extract(r"Loudness range:", r"LRA:\s*(-?\d+\.?\d*)")
    true_peak = extract(r"True peak:", r"Peak:\s*(-?\d+\.?\d*)")

    return {
        "lufs_i": lufs_i,
        "lufs_s_max": max(s_vals) if s_vals else None,
        "lufs_m_max": max(m_vals) if m_vals else None,
        "lra": lra,
        "true_peak_db": true_peak,
    }


def analyze_file(ffmpeg, ffprobe, path, cancel_token=None):
    meta = probe_file(ffprobe, path, cancel_token=cancel_token)
    basic = analyze_basic_stats(ffmpeg, path, meta["channels"], cancel_token=cancel_token)
    loud = analyze_loudness(ffmpeg, path, cancel_token=cancel_token)

    rms = basic["rms_db"]
    rms_l = rms[0] if len(rms) >= 1 else None
    rms_r = rms[1] if len(rms) >= 2 else None

    return {
        "format": path.suffix.upper().lstrip("."),
        "sample_rate": meta["sample_rate"],
        "bit_depth": meta["bit_depth"],
        "channels": meta["channels"],
        "duration": meta["duration"],
        "peak_db": basic["peak_db"],
        "rms_l_db": rms_l,
        "rms_r_db": rms_r,
        "dc_offset": basic["dc_offset"],
        "clip_count": basic["clip_count"],
        "lufs_i": loud["lufs_i"],
        "lufs_s_max": loud["lufs_s_max"],
        "lufs_m_max": loud["lufs_m_max"],
        "lra": loud["lra"],
        "true_peak_db": loud["true_peak_db"],
    }


def scan_folder(root_folders):
    files = []
    for folder in root_folders:
        folder = Path(folder)
        if folder.is_file():
            if folder.suffix.lower() in AUDIO_EXTS:
                files.append(folder)
            continue
        for dirpath, _, filenames in os.walk(folder):
            for name in filenames:
                if Path(name).suffix.lower() in AUDIO_EXTS:
                    files.append(Path(dirpath) / name)
    return sorted(set(files))


# Excel 导出：完整明细列，单位换行到第二行显示，避免窄列被截断
HEADERS = [
    "序号", "文件名", "采样率\n(Hz)", "位深", "声道数", "时长",
    "峰值\n(dBFS)", "左声道RMS\n(dBFS)", "右声道RMS\n(dBFS)",
    "平均响度\n(LUFS)", "短期最大响度\n(LUFS)", "瞬时最大响度\n(LUFS)",
    "响度范围\n(LU)", "真实峰值\n(dBTP)",
    "直流偏移", "是否削波", "备注",
]

# UI 表格：精简列，额外带上采样率/位深方便快速核对文件规格
TABLE_HEADERS = [
    "文件", "时长", "采样率", "位深", "峰值",
    "瞬时最大", "短期最大", "平均响度", "响度范围",
]


def build_excel_row(i, row):
    """把一条分析结果转换成 Excel 明细行（列顺序对应 HEADERS）"""
    if row.get("error"):
        return [
            i, row["name"], "", "", "", "",
            "", "", "", "", "", "", "", "", "", "", row["error"],
        ]

    d = row["data"]
    return [
        i, row["name"],
        d["sample_rate"] or "N/A",
        d["bit_depth"] if d["bit_depth"] else "N/A",
        d["channels"] or "N/A",
        fmt_seconds(d["duration"]),
        fmt_db(d["peak_db"]),
        fmt_db(d["rms_l_db"]),
        fmt_db(d["rms_r_db"]) if d["rms_r_db"] is not None else "-",
        fmt_db(d["lufs_i"]),
        fmt_db(d["lufs_s_max"]),
        fmt_db(d["lufs_m_max"]),
        fmt_db(d["lra"]),
        fmt_db(d["true_peak_db"]),
        round(d["dc_offset"], 6),
        "是" if d["clip_count"] > 0 else "否",
        "",
    ]


def build_table_row(row):
    """把一条分析结果转换成 UI 表格显示行（列顺序对应 TABLE_HEADERS）。
    row 可能是三种状态：
    - {name, dir}：仅列出，未检测 —— 除文件名列外全部留空
    - {name, dir, error}：检测/处理失败 —— 错误信息放在最后一列
    - {name, dir, data}：已检测/已处理 —— 各指标格式化后填入"""
    if row.get("error"):
        return [row["name"], "", "", "", "", "", "", "", row["error"]]

    d = row.get("data")
    if not d:
        return [row["name"]] + [""] * 8

    return [
        row["name"],
        fmt_seconds(d["duration"]),
        d["sample_rate"] or "N/A",
        d["bit_depth"] if d["bit_depth"] else "N/A",
        fmt_db(d["peak_db"], unit="dBFS"),
        fmt_db(d["lufs_m_max"], unit="LUFS"),
        fmt_db(d["lufs_s_max"], unit="LUFS"),
        fmt_db(d["lufs_i"], unit="LUFS"),
        fmt_db(d["lra"], unit="LU"),
    ]


# ---------------- 响度标准化（loudnorm 二次通过） ----------------

# loudnorm 内部工作在 192kHz 且线性模式对源 LRA 有上限；用 20 是 ffmpeg 允许的最大值，
# 让绝大多数音乐能走线性缩放而不是被压缩动态
LOUDNORM_TARGET_LRA = 20.0


def _clamp_bitrate(bit_rate, lo, hi, default):
    """把源码率截断到编码器可接受的范围；缺失时用 default"""
    if not bit_rate:
        return default
    return max(lo, min(hi, int(bit_rate)))


def _pick_output_codec(ext, meta):
    """按扩展名选择编码器，尽量与源文件参数保持一致。
    - PCM/WAV：按源位深选 pcm_s16le / pcm_s24le / pcm_s32le
    - FLAC：按源位深指定 sample_fmt（避免 24bit 源被降到 16bit）
    - 有损（MP3/AAC/OGG/OPUS/WMA）：`-b:a` 用源码率，缺失时回落到默认值"""
    ext = ext.lower()
    bit_depth = meta.get("bit_depth")
    codec_name = (meta.get("codec_name") or "").lower()
    bit_rate = meta.get("bit_rate")

    if ext == ".wav" or codec_name.startswith("pcm"):
        if bit_depth == 24:
            return ["-c:a", "pcm_s24le"]
        if bit_depth == 32:
            # 源为 32-bit float 时用 pcm_f32le 更贴近原始表示
            if (meta.get("sample_fmt") or "").startswith(("flt", "dbl")):
                return ["-c:a", "pcm_f32le"]
            return ["-c:a", "pcm_s32le"]
        if bit_depth == 8:
            return ["-c:a", "pcm_u8"]
        return ["-c:a", "pcm_s16le"]
    if ext == ".flac":
        # FLAC 只支持整数样本；把 24/32bit 源保为 s32 让编码器按最大精度存
        sample_fmt = "s16"
        if bit_depth and bit_depth >= 24:
            sample_fmt = "s32"
        return ["-c:a", "flac", "-sample_fmt", sample_fmt]
    if ext == ".mp3":
        br = _clamp_bitrate(bit_rate, 32_000, 320_000, 192_000)
        return ["-c:a", "libmp3lame", "-b:a", str(br)]
    if ext in (".m4a", ".aac"):
        br = _clamp_bitrate(bit_rate, 32_000, 512_000, 192_000)
        return ["-c:a", "aac", "-b:a", str(br)]
    if ext in (".ogg", ".oga"):
        br = _clamp_bitrate(bit_rate, 45_000, 500_000, 160_000)
        return ["-c:a", "libvorbis", "-b:a", str(br)]
    if ext == ".opus":
        br = _clamp_bitrate(bit_rate, 6_000, 510_000, 128_000)
        return ["-c:a", "libopus", "-b:a", str(br)]
    if ext in (".aiff", ".aif"):
        if bit_depth == 24:
            return ["-c:a", "pcm_s24be"]
        if bit_depth == 32:
            return ["-c:a", "pcm_s32be"]
        return ["-c:a", "pcm_s16be"]
    if ext == ".wma":
        br = _clamp_bitrate(bit_rate, 32_000, 384_000, 192_000)
        return ["-c:a", "wmav2", "-b:a", str(br)]
    return []


def _measure_loudnorm(ffmpeg, path, target_i, target_tp, target_lra, cancel_token=None):
    """loudnorm 第一遍：只测量，输出 JSON 到 stderr"""
    cmd = [
        ffmpeg, "-hide_banner", "-nostats", "-i", str(path),
        "-af",
        f"loudnorm=I={target_i}:TP={target_tp}:LRA={target_lra}:print_format=json",
        "-f", "null", "-",
    ]
    result = _run(cmd, cancel_token=cancel_token)
    text = result.stderr.decode(errors="ignore")
    start = text.rfind("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise RuntimeError(f"loudnorm 测量失败: {text[-300:]}")
    try:
        return json.loads(text[start:end + 1])
    except json.JSONDecodeError as e:
        raise RuntimeError(f"loudnorm 测量结果解析失败: {e}")


# 格式标准化：UI 里下拉框对应的固定选项集
FORMAT_LOSSLESS = {".wav", ".flac"}
FORMAT_LOSSY = {".mp3", ".m4a"}


def _codec_args_for_format(ext, bit_depth=None, bit_rate=None):
    """给定输出扩展名 + 位深/码率（按格式类型二选一），返回 ffmpeg codec 参数。
    比 _pick_output_codec 更明确 —— 前者是"跟随源"，这里是"用户指定"。"""
    ext = ext.lower()
    if ext == ".wav":
        if bit_depth == 24:
            return ["-c:a", "pcm_s24le"]
        if bit_depth == 32:
            return ["-c:a", "pcm_s32le"]
        return ["-c:a", "pcm_s16le"]
    if ext == ".flac":
        return ["-c:a", "flac", "-sample_fmt", "s32" if (bit_depth and bit_depth >= 24) else "s16"]
    if ext == ".mp3":
        br = int(bit_rate or 192_000)
        return ["-c:a", "libmp3lame", "-b:a", str(br)]
    if ext in (".m4a", ".aac"):
        br = int(bit_rate or 192_000)
        return ["-c:a", "aac", "-b:a", str(br)]
    return []


def _resolve_output_path(src_path, dst_dir, out_ext):
    """决定输出到 dst_dir 里的哪个文件名；扩展名不同时无需担心冲突，同扩展名+同目录 = 与源重名时加后缀避让"""
    src = Path(src_path)
    dst_dir = Path(dst_dir)
    dst = dst_dir / f"{src.stem}{out_ext}"
    # 如果输出会覆盖源文件本身（同扩展名 + 用户把输出目录选成了源目录），换个后缀避让
    try:
        same_as_src = dst.resolve() == src.resolve()
    except OSError:
        same_as_src = False
    if same_as_src:
        dst = dst_dir / f"{src.stem}_normalized{out_ext}"
    return dst


def process_file(
    ffmpeg, ffprobe, src_path, dst_dir,
    *,
    normalize_loudness=False,
    target_i=None, target_tp=None, tolerance_lu=None,
    measured_lufs=None, measured_tp=None,
    format_config=None,
    cancel_token=None,
):
    """把 src_path 处理后另存到 dst_dir。source 保持不变。
    - normalize_loudness=True 时走 loudnorm 二次通过（LUFS 已在容差且 TP 合规就跳过 loudnorm）
    - format_config is not None 时按 dict 指定的格式/采样率/位深或码率/声道 重新封装编码
    - 两者都为假：直接 copy 到目标目录
    - cancel_token：允许 UI 立即中断当前正在跑的 ffmpeg
    返回 (dst_path: Path, loudnorm_applied: bool, converted: bool)"""
    src = Path(src_path)
    src_meta = probe_file(ffprobe, src, cancel_token=cancel_token)

    # 决定输出扩展名与目标路径
    out_ext = (format_config or {}).get("ext") or src.suffix.lower()
    if not out_ext.startswith("."):
        out_ext = "." + out_ext
    Path(dst_dir).mkdir(parents=True, exist_ok=True)
    dst = _resolve_output_path(src, dst_dir, out_ext)

    # 判断是否需要跑 loudnorm
    lufs_ok = (
        measured_lufs is not None
        and measured_lufs != float("-inf")
        and target_i is not None
        and tolerance_lu is not None
        and abs(measured_lufs - target_i) <= tolerance_lu
    )
    tp_ok = (
        target_tp is None
        or measured_tp is None
        or measured_tp == float("-inf")
        or measured_tp <= target_tp
    )
    need_loudnorm = normalize_loudness and not (lufs_ok and tp_ok)

    # 都不做：直接复制
    if not need_loudnorm and not format_config:
        shutil.copy2(src, dst)
        return dst, False, False

    # 格式配置：优先用用户指定值，缺失时保留源参数
    if format_config:
        out_sr = format_config.get("sample_rate") or src_meta.get("sample_rate")
        out_ch = format_config.get("channels") or src_meta.get("channels")
        codec_args = _codec_args_for_format(
            out_ext,
            bit_depth=format_config.get("bit_depth"),
            bit_rate=format_config.get("bit_rate"),
        )
    else:
        out_sr = src_meta.get("sample_rate")
        out_ch = src_meta.get("channels")
        codec_args = _pick_output_codec(out_ext, src_meta)

    # loudnorm filter chain：两遍法（测量 + 应用），linear=true 保留动态。
    # 后面接 alimiter 做真峰保护：即便 loudnorm 因源 LRA > LOUDNORM_TARGET_LRA
    # 静默回退到 dynamic，或者 linear 结果的 inter-sample 峰值略超上限，
    # alimiter 也能兜底把峰压回目标 TP，不再放大动态压缩的量。
    af = None
    if need_loudnorm:
        stats = _measure_loudnorm(
            ffmpeg, src, target_i, target_tp, LOUDNORM_TARGET_LRA,
            cancel_token=cancel_token,
        )
        try:
            input_i = float(stats.get("input_i", "-inf"))
        except (TypeError, ValueError):
            input_i = float("-inf")
        # 素材几乎无声：跳过 loudnorm 步骤，但仍做格式转换（如果启用）
        if input_i <= -70.0 or input_i == float("-inf"):
            need_loudnorm = False
        else:
            # alimiter 只做峰值兜底：level=disabled/asc=off，避免自动电平把响度顶上去
            tp_linear = 10 ** (target_tp / 20.0)
            af = (
                f"loudnorm=I={target_i}:TP={target_tp}:LRA={LOUDNORM_TARGET_LRA}"
                f":measured_I={stats.get('input_i', 0)}"
                f":measured_TP={stats.get('input_tp', 0)}"
                f":measured_LRA={stats.get('input_lra', 0)}"
                f":measured_thresh={stats.get('input_thresh', -70)}"
                f":offset={stats.get('target_offset', 0)}"
                f":linear=true"
                f":print_format=summary,"
                f"alimiter=limit={tp_linear:.6f}:level=disabled:asc=0"
            )

    # 只需要 loudnorm 被跳过 + 也没有格式转换 = 直接 copy
    if not need_loudnorm and not format_config:
        shutil.copy2(src, dst)
        return dst, False, False

    cmd = [ffmpeg, "-y", "-hide_banner", "-v", "error", "-i", str(src)]
    if af:
        cmd += ["-af", af]
    if out_sr:
        cmd += ["-ar", str(out_sr)]
    if out_ch:
        cmd += ["-ac", str(out_ch)]
    cmd += codec_args
    cmd += [str(dst)]

    try:
        result = _run(cmd, cancel_token=cancel_token)
    except ProcCancelled:
        # 中断留下的半成品文件要清掉，避免下一轮误当已处理
        if dst.exists():
            try:
                dst.unlink()
            except OSError:
                pass
        raise
    if result.returncode != 0 or not dst.exists():
        if dst.exists():
            try:
                dst.unlink()
            except OSError:
                pass
        raise RuntimeError(
            f"处理失败: {result.stderr.decode(errors='ignore')[:300]}"
        )

    return dst, bool(af), bool(format_config)


def write_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "音频响度统计"
    ws.append(HEADERS)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
    # 表头行加高一档，让两行文字（名称 + 单位）都能显示完整
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    clip_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # 除文件名列（HEADERS 中第 2 列）外全部居中；文件名保持默认左对齐便于阅读
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    filename_col = 2

    for i, row in enumerate(rows, start=1):
        ws.append(build_excel_row(i, row))
        r = ws.max_row
        for cell in ws[r]:
            cell.alignment = left_align if cell.column == filename_col else center_align
        if row.get("error"):
            for cell in ws[r]:
                cell.fill = error_fill
        elif row["data"]["clip_count"] > 0:
            for cell in ws[r]:
                cell.fill = clip_fill

    widths = [6, 28, 10, 10, 8, 10, 11, 14, 14, 13, 15, 15, 11, 12, 10, 9, 30]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    wb.save(out_path)
